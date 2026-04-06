import flet as ft
from flet_color_pickers import ColorPicker
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from collections import defaultdict
import numpy as np
import re

# --- ORIGINAL EXTRACTION LOGIC ---
RAW_STRINGS = [
    "Stride length left front average in cm:",
    "Stride length right front average in cm:",
    "Stride length left hind average in cm:",
    "Stride length right hind average in cm:",
    "Overlap left average in cm:",
    "Overlap Right average in cm:",
    "Stride Width Front(L) average in cm:",
    "Stride Width Front(R) average in cm:",
    "Stride Width Hind(L) average in cm:",
    "Stride Width Hind(R) average in cm:"
]

RENAME_MAP = {
    "Stride Width Front(L) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Front(R) average in cm:": "Stride Width Front average in cm:",
    "Stride Width Hind(L) average in cm:": "Stride Width Hind average in cm:",
    "Stride Width Hind(R) average in cm:": "Stride Width Hind average in cm:"
}

hover_style = ft.ButtonStyle(mouse_cursor=ft.MouseCursor.CLICK)
hover_style_clear = ft.ButtonStyle(color=ft.Colors.GREY, mouse_cursor=ft.MouseCursor.CLICK)

def find_table_data(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    result_data = []
    THRESHOLD = 500

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_sheet_tag = sheet_name 

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Table ID"):
                    raw_table_id = cell.value.strip().replace("Table ID:", "").strip()
                    table_values = defaultdict(list)
                    table_values["Table ID"].append(raw_table_id)
                    table_values["Sheet_Source"].append(current_sheet_tag)

                    col_index = cell.column
                    row_index = cell.row

                    for r in range(row_index + 1, ws.max_row + 1):
                        label_cell = ws.cell(row=r, column=col_index)
                        value_cell = ws.cell(row=r, column=col_index + 1)
                        label = label_cell.value
                        value = value_cell.value

                        if label in RAW_STRINGS and isinstance(value, (int, float)):
                            output_label = RENAME_MAP.get(label, label)
                            if value > THRESHOLD:
                                value = value / 1000.0
                            table_values[output_label].append(value)

                    table_data = {}
                    for key, values in table_values.items():
                        if key in ["Table ID", "Sheet_Source"]:
                            table_data[key] = values[0]
                        else:
                            table_data[key] = float(np.mean(values)) if values else None

                    result_data.append(table_data)
    return result_data

# --- ADAPTED PROCESSING LOGIC ---
def process_file_adapted(input_file, id_to_color_map, output_dir=None, logger=print):
    logger(f"Extracting data from {os.path.basename(input_file)}...", ft.Colors.BLUE_400)
    data = find_table_data(input_file)
    if not data:
        raise ValueError(f"No 'Table ID' entries found in {os.path.basename(input_file)}.")

    w_groups = defaultdict(list)
    for entry in data:
        table_id = entry.get("Table ID", "")
        sheet_source = entry.get("Sheet_Source", "Unknown")
        found_tag = None

        parts = re.split(r'[_\s]+', table_id)
        for part in parts:
            if re.fullmatch(r"T\d+", part) or re.fullmatch(r"\d+W", part) or re.fullmatch(r"\d+\+\d+W", part):
                found_tag = part
                break
        
        if not found_tag:
             found_tag = sheet_source

        if found_tag:
            w_groups[found_tag].append(entry)
        else:
            w_groups["Unknown"].append(entry)

    original_name = os.path.splitext(os.path.basename(input_file))[0]
    timestamp = datetime.now().strftime("%H-%M-%S")
    
    # Change filename suffix based on whether colors are being applied
    file_suffix = "Colorized" if id_to_color_map else "Organized"
    
    target_directory = output_dir if output_dir else os.path.dirname(input_file)
    output_file = os.path.join(target_directory, f"{original_name}_{file_suffix}_{timestamp}.xlsx")

    logger("Building new workbook and applying styles...", ft.Colors.BLUE_200)
    wb = Workbook()
    wb.remove(wb.active)

    def sort_tags(tag):
        nums = re.findall(r'\d+', tag)
        if nums:
            return int(nums[0])
        return float('inf')

    def get_dynamic_font(hex_str):
        try:
            r, g, b = int(hex_str[2:4], 16), int(hex_str[4:6], 16), int(hex_str[6:8], 16)
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
            return Font(color="FF000000") if luminance > 0.5 else Font(color="FFFFFFFF")
        except:
            return Font(color="FF000000")

    for tag, group_data in sorted(w_groups.items(), key=lambda x: sort_tags(x[0])):
        clean_data = []
        for d in group_data:
            d_copy = d.copy()
            d_copy.pop("Sheet_Source", None)
            clean_data.append(d_copy)

        df = pd.DataFrame(clean_data)
        cols = ['Table ID'] + [c for c in df.columns if c != 'Table ID']
        df = df[cols]

        ws = wb.create_sheet(title=str(tag))

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = '0.000'

        headers = list(df.columns)
        try:
            table_id_col_index = headers.index("Table ID") + 1
        except ValueError:
            continue

        # Only iterate for styling if the user actually defined colors
        if id_to_color_map:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                table_id_cell = row[table_id_col_index - 1]
                table_id_val = str(table_id_cell.value).strip()

                matched_hex = None
                for user_id, hex_color in id_to_color_map.items():
                    if user_id in table_id_val:
                        matched_hex = hex_color
                        break

                if matched_hex:
                    fill = PatternFill(start_color=matched_hex, end_color=matched_hex, fill_type="solid")
                    font = get_dynamic_font(matched_hex)
                    for cell in row:
                        cell.fill = fill
                        cell.font = font

    wb.save(output_file)
    logger(f"Saved: {os.path.basename(output_file)}", ft.Colors.GREEN_400)
    return output_file


# --- FLET UI ---
def main(page: ft.Page):
    page.title = "Mouseframe Data Organizer"
    page.theme_mode = ft.ThemeMode.SYSTEM
    page.padding = 20
    page.scroll = ft.ScrollMode.AUTO

    groups = []
    selected_files = []
    selected_output_folder = [None] 

    # --- TERMINAL LOGGER ---
    log_view = ft.ListView(expand=False, height=150, spacing=4, auto_scroll=True)
    terminal_container = ft.Container(
        content=log_view,
        padding=10,
        bgcolor="#1e1e1e", 
        border_radius=8,
        border=ft.Border.all(1, ft.Colors.OUTLINE)
    )

    def log_msg(msg, text_color=ft.Colors.GREEN_400):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_view.controls.append(
            ft.Text(f"[{timestamp}] {msg}", color=text_color, font_family="monospace", size=13)
        )
        page.update()

    def clear_log(e):
        log_view.controls.clear()
        log_msg("System log cleared.", ft.Colors.GREY_400)
        page.update()

    log_header_row = ft.Row([
        ft.Text("System Log", size=14, weight=ft.FontWeight.BOLD),
        ft.TextButton("Clear Log", icon=ft.icons.Icons.CLEAR_ALL, on_click=clear_log, style=hover_style_clear)
    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)


    # --- THEME TOGGLE LOGIC ---
    def get_effective_theme():
        if page.theme_mode == ft.ThemeMode.SYSTEM:
            return ft.ThemeMode.DARK if page.platform_brightness == ft.Brightness.DARK else ft.ThemeMode.LIGHT
        return page.theme_mode

    def toggle_theme(e):
        current_theme = get_effective_theme()
        
        if current_theme == ft.ThemeMode.LIGHT:
            page.theme_mode = ft.ThemeMode.DARK
            theme_btn.icon = ft.icons.Icons.LIGHT_MODE
        else:
            page.theme_mode = ft.ThemeMode.LIGHT
            theme_btn.icon = ft.icons.Icons.DARK_MODE
            
        page.update()

    starting_theme = get_effective_theme()
    starting_icon = ft.icons.Icons.LIGHT_MODE if starting_theme == ft.ThemeMode.DARK else ft.icons.Icons.DARK_MODE

    theme_btn = ft.IconButton(
        icon=starting_icon, 
        on_click=toggle_theme, 
        tooltip="Toggle Light/Dark Mode", 
        style=hover_style
    )

    # --- TUTORIAL & HELP MODAL ---
    tutorial_md = """
### Welcome to the Mouseframe Group Organizer!

This tool allows you to automatically map colors and styles to specific animal IDs across your Excel datasets.

**Step 1: Select Input Data**
* Click **Browse Files** to select specific `.xlsx` files.
* *OR* click **Select Entire Folder** to bulk-load a whole directory of files.

**Step 2: Output Destination (Optional)**
* By default, the app saves the files in the same folder as the originals. You can override this by picking a specific folder.

**Step 3: Create Your Groups (Optional)**
* Type a group name (e.g., *Sick Male*).
* Click the colored square to pick the exact background color for this group in Excel.
* Type the Animal IDs separated by commas (e.g., `4023, 4036`).
* Click **Add Group**.
* *Note: If you skip this step, the app will simply extract and organize your data into clean sheets without adding any background colors.*

**Step 4: Process**
* Hit **Process Data**. The app will safely generate *copies* of your files (your originals are never overwritten!) and log the results in the terminal below.
    """

    help_dialog = ft.AlertDialog(
        title=ft.Row([ft.Icon(ft.icons.Icons.HELP), ft.Text("How to Use")]),
        content=ft.Container(
            content=ft.Column(
                controls=[ft.Markdown(tutorial_md, selectable=True, extension_set="gitHubWeb")],
                scroll=ft.ScrollMode.AUTO,
            ),
            width=500,  
            height=400, 
            padding=10
        ),
        actions=[ft.TextButton("Got it!", on_click=lambda e: setattr(help_dialog, 'open', False) or page.update(), style=hover_style)],
    )
    page.overlay.append(help_dialog)

    help_btn = ft.IconButton(
        icon=ft.icons.Icons.HELP_OUTLINE,
        tooltip="Help & Tutorial",
        on_click=lambda e: setattr(help_dialog, 'open', True) or page.update(),
        style=hover_style
    )

    header_row = ft.Row(
        [
            ft.Text("Mouseframe data organizer", size=28, weight=ft.FontWeight.BOLD),
            ft.Container(
                content=ft.Row([help_btn, theme_btn]), 
                padding=ft.Padding.only(right=15)
            )
        ], 
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN
    )

    # --- FILE & FOLDER PICKERS ---
    selected_file_text = ft.Text("No files selected", color=ft.Colors.RED, italic=True)
    output_folder_text = ft.Text("Default: Original file's folder", color=ft.Colors.GREY, italic=True)
    
    async def handle_pick_files(e):
        files = await ft.FilePicker().pick_files(allow_multiple=True, allowed_extensions=["xlsx", "xlsm"])
        selected_files.clear()
        
        if files:
            for f in files:
                selected_files.append(f.path)
            
            if len(selected_files) == 1:
                selected_file_text.value = os.path.basename(selected_files[0])
            else:
                selected_file_text.value = f"{len(selected_files)} files selected"
                
            selected_file_text.color = ft.Colors.GREEN
            selected_file_text.italic = False
            log_msg(f"Loaded {len(selected_files)} file(s) to process.", ft.Colors.WHITE)
        else:
            selected_file_text.value = "No files selected"
            selected_file_text.color = ft.Colors.RED
            selected_file_text.italic = True
            
        page.update()

    async def handle_pick_input_folder(e):
        folder = await ft.FilePicker().get_directory_path(dialog_title="Select Folder containing Excel files")
        
        if folder:
            valid_extensions = ('.xlsx', '.xlsm')
            found_files = [
                os.path.join(folder, f) for f in os.listdir(folder) 
                if f.lower().endswith(valid_extensions) and not f.startswith('~')
            ]
            
            selected_files.clear()
            
            if found_files:
                selected_files.extend(found_files)
                selected_file_text.value = f"{len(selected_files)} files loaded from folder"
                selected_file_text.color = ft.Colors.GREEN
                selected_file_text.italic = False
                log_msg(f"Loaded {len(selected_files)} file(s) from {os.path.basename(folder)}.", ft.Colors.WHITE)
            else:
                selected_file_text.value = "No valid Excel files found in folder"
                selected_file_text.color = ft.Colors.RED
                selected_file_text.italic = True
                log_msg(f"No .xlsx or .xlsm files found in {os.path.basename(folder)}.", ft.Colors.ORANGE_400)
                
        page.update()

    async def handle_pick_folder(e):
        folder = await ft.FilePicker().get_directory_path()
        if folder:
            selected_output_folder[0] = folder
            output_folder_text.value = folder
            output_folder_text.color = ft.Colors.GREEN
            output_folder_text.italic = False
            log_msg(f"Output folder set to: {folder}", ft.Colors.WHITE)
        page.update()

    # --- UI CONTROLS FOR GROUPS ---
    group_name_input = ft.TextField(label="Group Name (e.g. Sick Male)", expand=True)
    selected_color = {"hex": "#3366cc"} 
    
    color_preview = ft.Container(
        width=45, height=45, 
        bgcolor=selected_color["hex"], 
        border_radius=ft.BorderRadius.all(8),
        border=ft.Border.all(1, ft.Colors.BLACK_45),
        tooltip="Click to change color"
    )

    def on_color_change(e):
        selected_color["hex"] = e.data
        color_preview.bgcolor = e.data
        page.update()

    color_picker = ColorPicker(color=selected_color["hex"], on_color_change=on_color_change)

    def close_color_dialog(e):
        color_dialog.open = False
        page.update()

    def open_color_dialog(e):
        color_dialog.open = True
        page.update()

    color_dialog = ft.AlertDialog(
        title=ft.Text("Pick a Group Color"),
        content=color_picker,
        actions=[ft.TextButton("Done", on_click=close_color_dialog)],
    )
    page.overlay.append(color_dialog)

    groups_list = ft.Column(spacing=10)

    def delete_group(idx):
        removed = groups.pop(idx)
        log_msg(f"Removed group: {removed['name']}", ft.Colors.ORANGE_400)
        render_groups()

    def render_groups():
        groups_list.controls.clear()
        for i, grp in enumerate(groups):
            id_field = ft.TextField(
                label="Animal IDs (comma-separated, e.g., 4023, 4036)",
                value=grp["ids"],
                expand=True,
                on_change=lambda e, idx=i: groups[idx].update({"ids": e.control.value})
            )
            row = ft.Card(
                content=ft.Container(
                    padding=15,
                    content=ft.Row([
                        ft.Container(
                            width=30, height=30, bgcolor=grp["color"], 
                            border_radius=ft.BorderRadius.all(15), border=ft.Border.all(1, ft.Colors.BLACK_26)
                        ),
                        ft.Text(grp["name"], width=120, weight=ft.FontWeight.W_600),
                        id_field,
                        ft.IconButton(
                            icon=ft.icons.Icons.DELETE_OUTLINE, icon_color=ft.Colors.RED_400,
                            tooltip="Remove Group",
                            on_click=lambda e, idx=i: delete_group(idx), style=hover_style 
                        )
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
                )
            )
            groups_list.controls.append(row)
        page.update()

    def add_group(e):
        name = group_name_input.value.strip()
        if not name:
            page.snack_bar = ft.SnackBar(ft.Text("Group name cannot be empty!"), bgcolor=ft.Colors.RED)
            page.snack_bar.open = True
            page.update()
            return
            
        groups.append({"name": name, "color": selected_color["hex"], "ids": ""})
        log_msg(f"Created group '{name}' with color {selected_color['hex']}", ft.Colors.WHITE)
        group_name_input.value = ""
        render_groups()

    # --- MAIN PROCESS TRIGGER ---
    def process_data(e):
        if not selected_files:
            log_msg("Error: Attempted to process without selecting input files.", ft.Colors.RED_400)
            page.snack_bar = ft.SnackBar(ft.Text("Please select at least one Excel file first."), bgcolor=ft.Colors.RED)
            page.snack_bar.open = True
            page.update()
            return

        id_to_color_map = {}
        for grp in groups:
            hex_code = grp["color"].lstrip("#").upper()
            if len(hex_code) == 6:
                hex_code = "FF" + hex_code
            
            for animal_id in [p.strip() for p in grp["ids"].split(",") if p.strip()]:
                id_to_color_map[animal_id] = hex_code

        success_count = 0
        errors = []

        log_msg("--- STARTING BATCH PROCESS ---", ft.Colors.CYAN_400)

        for file_path in selected_files:
            try:
                process_file_adapted(file_path, id_to_color_map, selected_output_folder[0], log_msg)
                success_count += 1
            except Exception as ex:
                err_text = f"{os.path.basename(file_path)} failed: {ex}"
                errors.append(err_text)
                log_msg(err_text, ft.Colors.RED_400)

        if success_count == len(selected_files):
            msg = f"Success! Processed {success_count} file(s)."
            log_msg("--- BATCH COMPLETE: SUCCESS ---", ft.Colors.CYAN_400)
            page.snack_bar = ft.SnackBar(ft.Text(msg), bgcolor=ft.Colors.GREEN_700)
        elif success_count > 0:
            msg = f"Partial Success: Processed {success_count} files, but {len(errors)} failed."
            log_msg("--- BATCH COMPLETE: WITH ERRORS ---", ft.Colors.ORANGE_400)
            page.snack_bar = ft.SnackBar(ft.Text(msg), bgcolor=ft.Colors.ORANGE_700)
        else:
            log_msg("--- BATCH COMPLETE: ALL FAILED ---", ft.Colors.RED_400)
            page.snack_bar = ft.SnackBar(ft.Text(f"Error processing files: {errors[0]}"), bgcolor=ft.Colors.RED_700)
        
        page.snack_bar.open = True
        page.update()
    
    # Init system log
    log_msg("Application started. Waiting for user input...", ft.Colors.GREY_400)

    # --- RENDER DOM ---
    page.add(
        header_row,
        ft.Divider(),
        
        ft.Text("1. Select Target Excel File(s)", size=18, weight=ft.FontWeight.BOLD),
        ft.Row([
            ft.Button(
                "Browse Files", 
                icon=ft.icons.Icons.UPLOAD_FILE, 
                on_click=handle_pick_files, style=hover_style 
            ),
            ft.Text(" OR ", weight=ft.FontWeight.BOLD, color=ft.Colors.GREY_500),
            ft.Button(
                "Select Entire Folder", 
                icon=ft.icons.Icons.FOLDER, 
                on_click=handle_pick_input_folder, style=hover_style 
            ),
            selected_file_text
        ]),
        ft.Divider(),

        ft.Text("2. Select Output Folder (Optional)", size=18, weight=ft.FontWeight.BOLD),
        ft.Row([
            ft.Button(
                "Choose Folder", 
                icon=ft.icons.Icons.FOLDER_OPEN, 
                on_click=handle_pick_folder, style=hover_style 
            ),
            output_folder_text
        ]),
        ft.Divider(),

        # --- CHANGED HERE: Marked as Optional ---
        ft.Text("3. Create Groups & Assign Colors (Optional)", size=18, weight=ft.FontWeight.BOLD),
        ft.Row([
            group_name_input,
            ft.Text("Select Color:"),
            ft.GestureDetector(
                content=color_preview,
                on_tap=open_color_dialog, mouse_cursor='click'
            ),
            ft.Button("Add Group", icon=ft.icons.Icons.ADD, on_click=add_group, bgcolor=ft.Colors.BLUE, color=ft.Colors.WHITE, style=hover_style)
        ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER, margin=ft.Margin(0, 0, 20,0)),
        
        ft.Container(height=10),
        groups_list,
        ft.Divider(),

        ft.Text("4. Run Process", size=18, weight=ft.FontWeight.BOLD),
        ft.Button("Process Data", icon=ft.icons.Icons.PLAY_ARROW, on_click=process_data, bgcolor=ft.Colors.GREEN_600, color=ft.Colors.WHITE, height=50, style=hover_style),
        
        ft.Container(height=10),
        log_header_row,
        terminal_container
    )

ft.run(main)